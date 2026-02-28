"""Generate 3x/4x/5x Competitive Filing Landscape Excel for Scott.

Queries the local DB for all 3x/4x/5x fund filings and builds a single
Excel workbook with sheets: Summary, 3x, 4x, 5x, REX Only, Missing.

Output: reports/Competitive_Filing_Landscape_{date}.xlsx
"""
from __future__ import annotations

import re
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = PROJECT_ROOT / "data" / "etp_tracker.db"
REPORT_DIR = PROJECT_ROOT / "reports"
REPORT_DIR.mkdir(exist_ok=True)

TODAY = datetime.now().strftime("%Y-%m-%d")
XLSX_PATH = REPORT_DIR / f"Competitive_Filing_Landscape_{TODAY}.xlsx"

# ---------------------------------------------------------------------------
# Issuer mapping
# ---------------------------------------------------------------------------
ISSUER_MAP = {
    "ETF Opportunities Trust": "T-REX",
    "World Funds Trust": "T-REX",
    "Direxion Shares ETF Trust": "Direxion",
    "Direxion Funds": "Direxion",
    "ProShares Trust": "ProShares",
    "GraniteShares ETF Trust": "GraniteShares",
    "ETF Series Solutions": "Defiance",
    "Volatility Shares Trust": "Vol Shares",
    "Tidal Trust II": "LevMax",
    "Roundhill ETF Trust": "Roundhill",
    "Investment Managers Series Trust II": "Tradr",
    "Themes ETF Trust": "Lev Shares",
    "NEOS ETF Trust": "Kurv",
    "REX ETF Trust": "REX",
}

ISSUER_ORDER = [
    "T-REX", "ProShares", "Direxion", "GraniteShares", "Defiance",
    "Vol Shares", "Tradr", "Lev Shares", "LevMax", "Roundhill",
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
NAVY_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
GREEN_FILL = PatternFill(start_color="d5f5e3", end_color="d5f5e3", fill_type="solid")
BLUE_FILL = PatternFill(start_color="d6eaf8", end_color="d6eaf8", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="f2f3f4", end_color="f2f3f4", fill_type="solid")
REX_GREEN_FILL = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BOLD = Font(name="Calibri", bold=True, size=10)
BOLD_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NORMAL = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1a1a2e")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="555555")
KPI_NUM_FONT = Font(name="Calibri", bold=True, size=22, color="1a1a2e")
KPI_LABEL_FONT = Font(name="Calibri", size=9, color="777777")

THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------

def get_leverage(name: str) -> str | None:
    n = name.upper()
    if re.search(r"\b5[Xx]\b", n):
        return "5x"
    if re.search(r"\b4[Xx]\b|DAILY TARGET 4X", n):
        return "4x"
    if re.search(r"\b3[Xx]\b|DAILY TARGET 3X|3X1\b|\[MONTHLY 3X1\]|3X LONG|3X SHORT|3X INVERSE", n, re.IGNORECASE):
        return "3x"
    return None


def extract_underlier(name: str) -> str | None:
    n = name.strip()
    patterns = [
        r"T-REX\s+\d[Xx]\s+(?:LONG|INVERSE|SHORT)\s+(.+?)\s+DAILY",
        r"Direxion Daily\s+(.+?)\s+(?:Bull|Bear)\s+[345][Xx]",
        r"ProShares Daily Target \d[Xx]\s+(.+?)$",
        r"GraniteShares\s+\d[Xx]\s+(?:Long|Short|Inverse)\s+(.+?)\s+Daily",
        r"Defiance Daily Target \d[Xx]\s+(?:Long|Short|Inverse)\s+(.+?)\s+ETF",
        r"^[345][Xx]\s+(.+?)\s+ETF",
        r"LevMax\S*\s+(.+?)\s+\[",
        r"[Tt][Rr]adr\s+\d[Xx]\s+(?:Long|Short)\s+(.+?)\s+(?:Daily|Weekly|Monthly|Quarterly)",
        r"Leverage Shares\s+\d[Xx]\s+(?:Long|Short)\s+(.+?)\s+Daily",
        r"Roundhill\s+\d[Xx]\s+(.+?)\s+ETF",
    ]
    for pat in patterns:
        m = re.search(pat, n, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


def normalize_underlier(u: str) -> str:
    u = u.upper()
    u = u.replace("ALPHABET", "GOOGL")
    if u == "GOOG":
        u = "GOOGL"
    u = u.replace("BITCOIN", "BTC").replace("ETHER", "ETH")
    u = u.replace("BRK-B", "BRKB").replace("BRK.B", "BRKB")
    return u


def load_data():
    db = sqlite3.connect(str(DB_PATH))
    db.row_factory = sqlite3.Row
    rows = db.execute("""
        SELECT fs.fund_name, fs.status, fs.ticker, t.name as trust_name
        FROM fund_status fs
        JOIN trusts t ON fs.trust_id = t.id
        ORDER BY t.name, fs.fund_name
    """).fetchall()
    db.close()

    # matrix[leverage][underlier][issuer] = True
    matrices = {"3x": defaultdict(dict), "4x": defaultdict(dict), "5x": defaultdict(dict)}

    for r in rows:
        lev = get_leverage(r["fund_name"])
        if not lev:
            continue
        issuer = ISSUER_MAP.get(r["trust_name"], r["trust_name"])
        underlier = extract_underlier(r["fund_name"])
        if not underlier:
            continue
        underlier = normalize_underlier(underlier)
        matrices[lev][underlier][issuer] = True

    return {k: dict(v) for k, v in matrices.items()}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _active_issuers(matrix):
    """Get ordered list of issuers that appear in this matrix."""
    seen = set()
    for issuers in matrix.values():
        seen.update(issuers.keys())
    return [i for i in ISSUER_ORDER if i in seen]


def _write_matrix_sheet(wb, title, matrix):
    """Write a leverage matrix sheet. Columns: Underlier, then one per issuer.
    Cells show X where an issuer has filed."""
    ws = wb.create_sheet(title)
    issuers = _active_issuers(matrix)
    headers = ["Underlier"] + issuers
    ncols = len(headers)

    # Find T-REX column index (1-based)
    trex_col = issuers.index("T-REX") + 2 if "T-REX" in issuers else None

    # Header row
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, ncols)

    # If T-REX column, give it a distinct header color
    if trex_col:
        cell = ws.cell(row=1, column=trex_col)
        cell.fill = REX_GREEN_FILL
        cell.font = BOLD_WHITE

    # Data rows
    for ri, underlier in enumerate(sorted(matrix.keys()), 2):
        issuers_map = matrix[underlier]

        cell = ws.cell(row=ri, column=1, value=underlier)
        cell.font = BOLD
        cell.border = THIN_BORDER
        cell.alignment = LEFT

        for ci, iss in enumerate(issuers, 2):
            cell = ws.cell(row=ri, column=ci)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if iss in issuers_map:
                cell.value = "X"
                cell.font = BOLD
                if ci == trex_col:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = BLUE_FILL
            else:
                cell.font = NORMAL

        # Alternate row shading for empty cells
        if ri % 2 == 0:
            for ci in range(1, ncols + 1):
                cell = ws.cell(row=ri, column=ci)
                if cell.fill == PatternFill() or cell.fill is None:
                    pass  # keep colored cells as-is

    # Column widths
    _auto_width(ws, 1, 28)
    for ci in range(2, ncols + 1):
        _auto_width(ws, ci, 14)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "B2"
    return ws


def _write_summary_sheet(wb, matrices):
    """Summary sheet: issuer scorecard + counts."""
    ws = wb.active
    ws.title = "Summary"

    report_date = datetime.now().strftime("%B %d, %Y")

    # Title
    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value="3x / 4x / 5x Competitive Filing Landscape")
    cell.font = TITLE_FONT
    cell.alignment = LEFT

    ws.merge_cells("A2:G2")
    cell = ws.cell(row=2, column=1, value=f"Source: SEC EDGAR  |  {report_date}")
    cell.font = SUBTITLE_FONT
    cell.alignment = LEFT

    # KPI row
    row = 4
    kpis = [
        ("3x Names", len(matrices["3x"])),
        ("4x Names", len(matrices["4x"])),
        ("5x Names", len(matrices["5x"])),
    ]
    for ci, (label, val) in enumerate(kpis):
        col = 1 + ci * 2
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = KPI_NUM_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        cell2 = ws.cell(row=row + 1, column=col, value=label)
        cell2.font = KPI_LABEL_FONT
        cell2.alignment = CENTER
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)

    # Issuer scorecard table
    row = 7
    ws.cell(row=row, column=1, value="Issuer Scorecard").font = Font(
        name="Calibri", bold=True, size=12, color="1a1a2e")
    row = 8
    headers = ["Issuer", "3x", "4x", "5x", "Total", "Exclusive"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    _style_header(ws, row, len(headers))

    # Gather all issuers
    all_issuers = set()
    for m in matrices.values():
        for issuers in m.values():
            all_issuers.update(issuers.keys())
    ordered = [i for i in ISSUER_ORDER if i in all_issuers]

    for ri, iss in enumerate(ordered, row + 1):
        c3 = sum(1 for u, i in matrices["3x"].items() if iss in i)
        c4 = sum(1 for u, i in matrices["4x"].items() if iss in i)
        c5 = sum(1 for u, i in matrices["5x"].items() if iss in i)
        total = c3 + c4 + c5

        # Exclusive = names where this issuer is the ONLY filer (across all leverage)
        excl = 0
        for lev in ("3x", "4x", "5x"):
            for u, issuers_map in matrices[lev].items():
                if iss in issuers_map and len(issuers_map) == 1:
                    excl += 1

        vals = [iss, c3, c4, c5, total, excl]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if ci > 1 else LEFT
            cell.font = NORMAL

        # Highlight T-REX row
        if iss == "T-REX":
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = GREEN_FILL
                ws.cell(row=ri, column=ci).font = BOLD

    last_issuer_row = row + len(ordered)

    # Most contested names
    contest_row = last_issuer_row + 2
    ws.cell(row=contest_row, column=1, value="Most Contested Names (by # of issuers)").font = Font(
        name="Calibri", bold=True, size=12, color="1a1a2e")
    contest_row += 1
    c_headers = ["Name", "Leverage", "# Issuers", "Issuers"]
    for ci, h in enumerate(c_headers, 1):
        ws.cell(row=contest_row, column=ci, value=h)
    _style_header(ws, contest_row, len(c_headers))

    # Combine all matrices, sort by # issuers desc
    all_entries = []
    for lev, matrix in matrices.items():
        for u, issuers_map in matrix.items():
            all_entries.append((u, lev, len(issuers_map), ", ".join(
                sorted(issuers_map.keys(),
                       key=lambda x: ISSUER_ORDER.index(x) if x in ISSUER_ORDER else 99))))
    all_entries.sort(key=lambda x: (-x[2], x[1], x[0]))

    for ri, (u, lev, count, who) in enumerate(all_entries[:25], contest_row + 1):
        vals = [u, lev, count, who]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if ci in (2, 3) else LEFT
            cell.font = NORMAL

    # Column widths
    _auto_width(ws, 1, 22)
    _auto_width(ws, 2, 10)
    _auto_width(ws, 3, 10)
    _auto_width(ws, 4, 12)
    _auto_width(ws, 5, 12)
    _auto_width(ws, 6, 12)
    # For contest section, col 4 needs to be wider
    # (openpyxl uses max width per column, so set it wide enough)
    _auto_width(ws, 4, 60)


def _write_rex_only_sheet(wb, matrices):
    """Names where T-REX is the only issuer."""
    ws = wb.create_sheet("REX Only")
    headers = ["Name", "Leverage"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, len(headers))

    ri = 2
    for lev in ("3x", "4x", "5x"):
        for u, issuers_map in sorted(matrices[lev].items()):
            if "T-REX" in issuers_map and len(issuers_map) == 1:
                ws.cell(row=ri, column=1, value=u).font = BOLD
                ws.cell(row=ri, column=1).border = THIN_BORDER
                ws.cell(row=ri, column=1).alignment = LEFT
                ws.cell(row=ri, column=2, value=lev).font = NORMAL
                ws.cell(row=ri, column=2).border = THIN_BORDER
                ws.cell(row=ri, column=2).alignment = CENTER
                # Green highlight
                for ci in (1, 2):
                    ws.cell(row=ri, column=ci).fill = GREEN_FILL
                ri += 1

    _auto_width(ws, 1, 28)
    _auto_width(ws, 2, 12)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _write_missing_sheet(wb, matrices):
    """Names where competitors have filed but T-REX has NOT."""
    ws = wb.create_sheet("Missing")
    headers = ["Name", "Leverage", "# Competitors", "Who Has Filed"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, len(headers))

    entries = []
    for lev in ("3x", "4x", "5x"):
        for u, issuers_map in matrices[lev].items():
            if "T-REX" not in issuers_map:
                who = ", ".join(sorted(issuers_map.keys(),
                    key=lambda x: ISSUER_ORDER.index(x) if x in ISSUER_ORDER else 99))
                entries.append((u, lev, len(issuers_map), who))
    entries.sort(key=lambda x: (-x[2], x[1], x[0]))

    for ri, (u, lev, count, who) in enumerate(entries, 2):
        ws.cell(row=ri, column=1, value=u).font = BOLD
        ws.cell(row=ri, column=1).border = THIN_BORDER
        ws.cell(row=ri, column=1).alignment = LEFT
        ws.cell(row=ri, column=2, value=lev).font = NORMAL
        ws.cell(row=ri, column=2).border = THIN_BORDER
        ws.cell(row=ri, column=2).alignment = CENTER
        ws.cell(row=ri, column=3, value=count).font = NORMAL
        ws.cell(row=ri, column=3).border = THIN_BORDER
        ws.cell(row=ri, column=3).alignment = CENTER
        ws.cell(row=ri, column=4, value=who).font = NORMAL
        ws.cell(row=ri, column=4).border = THIN_BORDER
        ws.cell(row=ri, column=4).alignment = LEFT

    _auto_width(ws, 1, 28)
    _auto_width(ws, 2, 12)
    _auto_width(ws, 3, 16)
    _auto_width(ws, 4, 60)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading data...")
    matrices = load_data()
    for lev in ("3x", "4x", "5x"):
        issuers = set()
        for i in matrices[lev].values():
            issuers.update(i.keys())
        print(f"  {lev}: {len(matrices[lev])} names, {len(issuers)} issuers")

    wb = Workbook()

    print("Writing Summary...")
    _write_summary_sheet(wb, matrices)

    print("Writing 3x...")
    _write_matrix_sheet(wb, "3x", matrices["3x"])

    print("Writing 4x...")
    _write_matrix_sheet(wb, "4x", matrices["4x"])

    print("Writing 5x...")
    _write_matrix_sheet(wb, "5x", matrices["5x"])

    print("Writing REX Only...")
    _write_rex_only_sheet(wb, matrices)

    print("Writing Missing...")
    _write_missing_sheet(wb, matrices)

    wb.save(str(XLSX_PATH))
    print(f"\nSaved: {XLSX_PATH}")


if __name__ == "__main__":
    main()
